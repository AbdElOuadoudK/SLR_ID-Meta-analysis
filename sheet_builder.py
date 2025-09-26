from __future__ import annotations

import os
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from output_paths import resolve_csv_dir

# Single-sheet column layout (merged table)
COLUMNS = [
    "mode", "paperId", "title", "publication_date", "year",
    "publication_types", "fields_of_study",
    "influential_citation_count", "citation_count",
    "abstract", "external_ids", "doi", "is_open_access", "open_access_pdf_url",
    "journal_pages_range", "pages_total", "references_pages", "references_count", "_max_cited_year",
    "authors_hindex_list", "mean_author_hindex",
]

# Helpers
def _write_headers(ws, headers):
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)


def build_template(path: Path, params: dict):
    """
    Build an .xlsx with a single sheet 'rawdata' that contains the raw input
    columns. This preserves the original semantics but keeps everything in one
    table.
    """
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws = wb.create_sheet("rawdata", 0)
    _write_headers(ws, COLUMNS)
    ws.freeze_panes = "A2"

    # Column index map
    idx = {h: i for i, h in enumerate(COLUMNS, start=1)}

    def col(h): return get_column_letter(idx[h])

    # Open-ended ranges for ARRAYFORMULA-like formulas (keeps original strings)
    def rng(h): return f"{col(h)}2:{col(h)}"

    # Row where formulas are placed (row 2 using ARRAYFORMULA semantics)
    r = 2

    # ---------------------------
    # Metric formulas (written into the same sheet)
    # ---------------------------

    # paperId: copy from same sheet (keeps row alignment)
    ws[f"{col('paperId')}{r}"] = f"=ARRAYFORMULA(IF(LEN({rng('paperId')})=0, \"\", {rng('paperId')}))"
    ws[f"{col('paperId')}{r}"].alignment = Alignment(wrapText=True)

    # Save template
    wb.save(path)


def build_template_with_data(outdir: str, params: dict, csv_dir: Optional[str] = None):
    """
    Create the single-sheet template, then populate 'rawdata' with extracted raw data
    from extracted_dataset.xlsx when available, and save the final workbook as metrics_with_formulas_single_sheet.xlsx.
    """
    base = Path(outdir)
    target_dir = resolve_csv_dir(base, csv_dir)
    template_path = target_dir / "metrics_template_single_sheet.xlsx"
    build_template(template_path, params)

    data_path = target_dir / "extracted_dataset.xlsx"
    if not os.path.exists(data_path):
        # Produce the template even if no extracted data is present yet.
        return

    df = pd.read_excel(data_path, dtype=str).fillna("")
    wb = load_workbook(template_path)
    ws = wb["rawdata"]

    headers = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    start = 2

    for i, row in df.iterrows():
        r = start + i
        for h, val in row.items():
            if h in col_idx:
                ws.cell(row=r, column=col_idx[h], value=("" if pd.isna(val) else val))

    wb.save(target_dir / "paperset.xlsx")
